import csv
import zipfile

from openpyxl import load_workbook
from selenium import webdriver
from selene import browser
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from os_path.os_path_scripts import tmp, resources
import os.path
import requests
from pypdf import PdfReader
import xlrd
import time
from zipfile import ZipFile

PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))

def test_open_csv():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    csv_file = os.path.join(PROJECT_ROOT_PATH, resources, 'eggs.csv')
    with open(csv_file, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open('resources/eggs.csv') as csvfile:
        csvreader = csv.reader(csvfile)
        check_list = []
        for row in csvreader:
            check_list.append(row)
        assert check_list[0] == ['Anna', 'Pavel', 'Peter']
        assert check_list[1] == ['Alex', 'Serj', 'Yana']

def test_open_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    pdf_file = os.path.join(PROJECT_ROOT_PATH, resources, 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_file)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


def test_open_xls():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xls_file = os.path.join(PROJECT_ROOT_PATH, resources, 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(xls_file)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))
    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    assert sheet.ncols == 8

def test_open_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xlsx_file = os.path.join(PROJECT_ROOT_PATH, 'resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'

def test_open_file_with_browser():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": os.path.join(PROJECT_ROOT_PATH, 'tmp'),
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    browser.config.driver = driver

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(3)
    file_element = os.path.join(tmp, 'pytest-main.zip')
    size_file = os.path.getsize(file_element)
    assert size_file == 1565741

def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    logo = os.path.join(PROJECT_ROOT_PATH, 'tmp', 'selenium_logo_square_green.png')
    r = requests.get(url)
    with open(logo, 'wb') as file:
            file.write(r.content)

    size = os.path.getsize(logo)

    assert size == 30803